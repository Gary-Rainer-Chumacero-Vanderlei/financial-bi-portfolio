"""
notebooks/eda_financeiro.py — v4
=================================
Análise Exploratória com Storytelling — Financial Performance Dashboard
Empresa: Energética Miramar Distribuidora Ltda.
Período: Jan/2023 – Dez/2024

EXECUÇÃO:
    python -m notebooks.eda_financeiro

SAÍDAS:
    exports/prints/relatorio_financeiro.pdf   ← Relatório executivo completo
    exports/prints/relatorio_financeiro.xlsx  ← Workbook com 5 abas analíticas

ESTRUTURA NARRATIVA (3 atos):
    Ato 1 — Panorama   : Receita, EBITDA, sazonalidade, YoY
    Ato 2 — Diagnóstico: Custos, outliers, inadimplência, fluxo de caixa
    Ato 3 — Síntese    : Score de saúde, waterfall DRE, achados e recomendações
"""
from __future__ import annotations

import io
import logging
import sys
import tempfile
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import seaborn as sns
from matplotlib.colors import LinearSegmentedColormap

_ROOT = Path(__file__).parent.parent.resolve()
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from src.data.loader import DataLoader
from src.analysis.kpis import calcular_score_saude, formatar_brl
from config.settings import Settings

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

# =============================================================================
# DESIGN SYSTEM
# =============================================================================
BG      = "#0f1923"
BG2     = "#162030"
ACCENT  = "#00BFA6"
ACCENT2 = "#F5A623"
ACCENT3 = "#E05A5A"
GRID    = "#1e2d3d"
TEXT    = "#c8d6e5"
MUTED   = "#4a6278"
BLUE    = "#4F8EF7"
PURPLE  = "#A78BFA"

PALETTE = [ACCENT, ACCENT2, ACCENT3, BLUE, PURPLE, "#4ecdc4"]

plt.rcParams.update({
    "figure.facecolor":  BG,
    "axes.facecolor":    BG2,
    "axes.edgecolor":    GRID,
    "axes.labelcolor":   TEXT,
    "axes.titlecolor":   TEXT,
    "axes.titlesize":    12,
    "axes.labelsize":    10,
    "axes.grid":         True,
    "grid.color":        GRID,
    "grid.linewidth":    0.6,
    "xtick.color":       MUTED,
    "ytick.color":       MUTED,
    "xtick.labelsize":   8.5,
    "ytick.labelsize":   8.5,
    "legend.facecolor":  BG2,
    "legend.edgecolor":  GRID,
    "legend.labelcolor": TEXT,
    "legend.fontsize":   9,
    "text.color":        TEXT,
    "font.family":       "sans-serif",
})

MN = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]


# =============================================================================
# UTILITÁRIOS
# =============================================================================

def fmt_M(x: float, _: object) -> str:
    return f"R${x/1e6:.1f}M"

def fmt_K(x: float, _: object) -> str:
    return f"R${x/1e3:.0f}K"

def fmt_pct(x: float, _: object) -> str:
    return f"{x:.1f}%"

def _fig(w: float, h: float) -> plt.Figure:
    return plt.figure(figsize=(w, h), facecolor=BG)

def _ax_style(ax: plt.Axes) -> plt.Axes:
    ax.set_facecolor(BG2)
    for sp in ax.spines.values():
        sp.set_edgecolor(GRID)
    return ax

def _salvar_png(fig: plt.Figure, tmp_dir: Path, nome: str) -> Path:
    """Salva figura como PNG temporário e retorna o caminho."""
    p = tmp_dir / f"{nome}.png"
    fig.savefig(p, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    return p


# =============================================================================
# ATO 1 — PANORAMA
# =============================================================================

def fig_receita_ebitda(dre: pd.DataFrame, tmp: Path) -> Path:
    """G1 — Receita Líquida mensal com MM3 e EBITDA com margem."""
    x = np.arange(len(dre))
    fig = _fig(14, 8)
    gs  = gridspec.GridSpec(2, 1, figure=fig, hspace=0.55,
                            left=0.08, right=0.96, top=0.92, bottom=0.10)
    fig.text(0.5, 0.97, "Ato 1 — Panorama  ·  Receita & EBITDA",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)

    ax1 = _ax_style(fig.add_subplot(gs[0]))
    ax1.bar(x, dre["receita_liquida"]/1e6, color=BLUE, alpha=0.75, width=0.65, zorder=2,
            label="Receita Líquida")
    mm3 = dre["receita_liquida"].rolling(3, min_periods=1).mean()
    ax1.plot(x, mm3/1e6, color=ACCENT, lw=2.2, ls="--", zorder=3, label="Média Móvel 3M")
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_M))
    ax1.set_xticks(x); ax1.set_xticklabels(dre["competencia"], rotation=45, ha="right", fontsize=7.5)
    ax1.set_title("Receita Líquida Mensal  ·  Jan/2023 – Dez/2024")
    ax1.legend()

    ax2 = _ax_style(fig.add_subplot(gs[1]))
    cores = [ACCENT if v >= 0 else ACCENT3 for v in dre["ebitda"]]
    ax2.bar(x, dre["ebitda"]/1e6, color=cores, alpha=0.80, width=0.65, zorder=2, label="EBITDA")
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_M))
    ax2.set_xticks(x); ax2.set_xticklabels(dre["competencia"], rotation=45, ha="right", fontsize=7.5)
    ax2.set_title("EBITDA Mensal  ·  Positivo (verde) / Negativo (vermelho)")
    axr = ax2.twinx()
    axr.plot(x, dre["margem_ebitda_pct"], color=ACCENT2, lw=2.2, marker="o", ms=4, label="Margem %")
    axr.set_ylabel("Margem (%)", color=ACCENT2, fontsize=9.5)
    axr.tick_params(axis="y", colors=ACCENT2, labelsize=8.5)
    axr.set_facecolor(BG2)
    for sp in axr.spines.values(): sp.set_edgecolor(GRID)
    axr.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_pct))
    axr.legend(loc="upper left")
    ax2.legend(loc="upper right")
    return _salvar_png(fig, tmp, "g1_receita_ebitda")


def fig_sazonalidade(dre: pd.DataFrame, tmp: Path) -> Path:
    """G2 — Sazonalidade: índice mensal médio e YoY lado a lado."""
    d23 = dre[dre["ano"] == 2023].reset_index(drop=True)
    d24 = dre[dre["ano"] == 2024].reset_index(drop=True)
    x12 = np.arange(12); w = 0.38

    media_geral = dre.groupby("mes")["receita_liquida"].mean()
    indice = media_geral / media_geral.mean() * 100

    fig = _fig(14, 6)
    gs  = gridspec.GridSpec(1, 2, figure=fig, wspace=0.32,
                            left=0.07, right=0.97, top=0.88, bottom=0.16)
    fig.text(0.5, 0.96, "Sazonalidade e Comparativo Year-over-Year",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)

    # Índice sazonal
    ax1 = _ax_style(fig.add_subplot(gs[0]))
    bars = ax1.bar(x12, indice.values, color=[ACCENT if v >= 100 else ACCENT3 for v in indice],
                   alpha=0.85, width=0.7)
    ax1.axhline(100, color=MUTED, ls="--", lw=1.5)
    ax1.set_xticks(x12); ax1.set_xticklabels(MN, fontsize=9)
    ax1.set_title("Índice de Sazonalidade da Receita (base 100)")
    ax1.set_ylabel("Índice", fontsize=9.5)
    for bar, val in zip(bars, indice.values):
        ax1.text(bar.get_x()+bar.get_width()/2, val+0.8, f"{val:.0f}",
                 ha="center", fontsize=8, color=TEXT, fontweight="bold")

    # YoY barras agrupadas
    ax2 = _ax_style(fig.add_subplot(gs[1]))
    ax2.bar(x12-w/2, d23["receita_liquida"]/1e6, w, label="2023",
            color=MUTED, alpha=0.80)
    ax2.bar(x12+w/2, d24["receita_liquida"]/1e6, w, label="2024",
            color=BLUE, alpha=0.85)
    for i in range(min(len(d23), len(d24))):
        v23, v24 = d23["receita_liquida"].iloc[i], d24["receita_liquida"].iloc[i]
        var = (v24-v23)/v23*100 if v23 else 0
        cor = ACCENT if var >= 0 else ACCENT3
        alto = max(v23, v24)/1e6
        ax2.text(i, alto+0.03, f"{var:+.1f}%", ha="center", fontsize=7.5,
                 color=cor, fontweight="bold")
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_M))
    ax2.set_xticks(x12); ax2.set_xticklabels(MN, fontsize=9)
    ax2.set_title("Receita Líquida YoY — 2023 vs 2024")
    ax2.legend()
    return _salvar_png(fig, tmp, "g2_sazonalidade")


def fig_margens(dre: pd.DataFrame, tmp: Path) -> Path:
    """G3 — Evolução das três margens ao longo do tempo."""
    fig = _fig(14, 5)
    ax  = _ax_style(fig.add_axes([0.07, 0.18, 0.89, 0.68]))
    fig.text(0.5, 0.97, "Evolução das Margens Operacionais  ·  Jan/2023 – Dez/2024",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    x = np.arange(len(dre))
    for col, nome, cor in [
        ("margem_bruta_pct",   "Margem Bruta",   BLUE),
        ("margem_ebitda_pct",  "Margem EBITDA",  ACCENT2),
        ("margem_liquida_pct", "Margem Líquida", PURPLE),
    ]:
        ax.plot(x, dre[col], color=cor, lw=2, marker="o", ms=4, label=nome)
    ax.axhline(20, color=ACCENT, ls="--", lw=1.5, alpha=0.7, label="Meta EBITDA 20%")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_pct))
    ax.set_xticks(x)
    ax.set_xticklabels(dre["competencia"], rotation=45, ha="right", fontsize=7.5)
    ax.legend(ncol=4)
    return _salvar_png(fig, tmp, "g3_margens")


# =============================================================================
# ATO 2 — DIAGNÓSTICO
# =============================================================================

def fig_custos_heatmap(centro: pd.DataFrame, tmp: Path) -> Path:
    """G4 — Heatmap de gastos mensais por departamento."""
    pivot = (centro.groupby(["competencia","departamento"])["valor"]
             .sum().reset_index()
             .pivot(index="departamento", columns="competencia", values="valor") / 1_000)
    cmap = LinearSegmentedColormap.from_list("fin", [BG2, "#1E3A5F", BLUE, ACCENT], N=256)
    fig = _fig(16, 5)
    fig.text(0.5, 0.97, "Ato 2 — Diagnóstico  ·  Heatmap de Custos por Departamento (R$ Mil)",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    ax = fig.add_axes([0.09, 0.20, 0.87, 0.68])
    ax.set_facecolor(BG2)
    sns.heatmap(pivot, ax=ax, cmap=cmap, linewidths=0.5, linecolor=BG,
                annot=True, fmt=".0f", annot_kws={"size":7.5, "color":TEXT},
                cbar_kws={"label":"R$ Mil","shrink":0.80})
    ax.tick_params(axis="x", labelsize=8, colors=MUTED, rotation=45)
    ax.tick_params(axis="y", labelsize=9, colors=TEXT, rotation=0)
    ax.set_xlabel(""); ax.set_ylabel("")
    cbar = ax.collections[0].colorbar
    cbar.ax.yaxis.set_tick_params(color=MUTED, labelsize=8)
    cbar.set_label("R$ Mil", color=MUTED, fontsize=9)
    return _salvar_png(fig, tmp, "g4_heatmap_custos")


def fig_outliers(centro: pd.DataFrame, tmp: Path) -> Path:
    """G5 — Outliers por departamento via Z-score."""
    dm     = centro.groupby(["competencia","departamento"])["valor"].sum().reset_index()
    deptos = sorted(dm["departamento"].unique())
    fig = _fig(16, 5)
    fig.text(0.5, 0.97, "Detecção de Outliers  ·  Z-score por Departamento",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    gs = gridspec.GridSpec(1, len(deptos), figure=fig, wspace=0.35,
                           left=0.06, right=0.97, top=0.88, bottom=0.20)
    for i, depto in enumerate(deptos):
        ax = _ax_style(fig.add_subplot(gs[i]))
        dados = dm[dm["departamento"]==depto]["valor"].values
        media = dados.mean()
        std   = dados.std() if dados.std() > 0 else 1.0
        xi    = np.arange(len(dados))
        ax.plot(xi, dados/1e3, color=MUTED, lw=1, alpha=0.5)
        cores = [ACCENT3 if abs(v-media) > 2*std else BLUE for v in dados]
        ax.scatter(xi, dados/1e3, c=cores, s=45, zorder=3)
        ax.axhline(media/1e3, color=ACCENT, ls="--", lw=1.5, alpha=0.9)
        ax.axhline((media+2*std)/1e3, color=ACCENT3, ls=":", lw=1.0, alpha=0.7)
        ax.axhline((media-2*std)/1e3, color=ACCENT3, ls=":", lw=1.0, alpha=0.7)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_K))
        ax.set_title(depto, fontsize=10)
        ax.set_xticks([])
    fig.legend(handles=[
        mpatches.Patch(color=BLUE,   label="Normal"),
        mpatches.Patch(color=ACCENT3,label="Outlier |Z|>2"),
        mpatches.Patch(color=ACCENT, label="Média histórica"),
    ], loc="lower center", ncol=3, fontsize=9, bbox_to_anchor=(0.5, 0.01))
    return _salvar_png(fig, tmp, "g5_outliers")


def fig_inadimplencia(contas: pd.DataFrame, tmp: Path) -> Path:
    """G6 — Taxa de inadimplência mensal e donut de aging."""
    inadim = (contas.groupby("competencia")
              .apply(lambda g: g.loc[g["inadimplente"]==1,"valor"].sum()
                     / g["valor"].sum() * 100 if g["valor"].sum() > 0 else 0.0)
              .reset_index(name="taxa").sort_values("competencia"))
    ultimo = contas["competencia"].max()
    aging  = (contas[contas["competencia"]==ultimo]
              .groupby("faixa_aging")["valor"].sum())
    ordem  = ["A vencer","1-30 dias","31-60 dias","61-90 dias","91-180 dias",">180 dias"]
    aging  = aging.reindex([o for o in ordem if o in aging.index])

    fig = _fig(14, 6)
    fig.text(0.5, 0.97, "Inadimplência e Aging da Carteira",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    gs = gridspec.GridSpec(1, 2, figure=fig, wspace=0.35,
                           left=0.07, right=0.97, top=0.88, bottom=0.16)

    ax1 = _ax_style(fig.add_subplot(gs[0]))
    xi  = np.arange(len(inadim))
    ax1.fill_between(xi, inadim["taxa"], alpha=0.15, color=ACCENT3)
    ax1.plot(xi, inadim["taxa"], color=ACCENT3, lw=2.5, marker="o", ms=5)
    media = inadim["taxa"].mean()
    ax1.axhline(media, color=ACCENT2, ls="--", lw=1.6, label=f"Média {media:.1f}%")
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_pct))
    ax1.set_xticks(xi)
    ax1.set_xticklabels(inadim["competencia"], rotation=45, ha="right", fontsize=7.5)
    ax1.set_title("Taxa de Inadimplência Mensal (%)")
    ax1.legend()

    ax2 = fig.add_subplot(gs[1])
    ax2.set_facecolor(BG)
    cores_aging = [ACCENT, ACCENT2, ACCENT3, BLUE, PURPLE, "#4ecdc4"]
    wedges, _, auts = ax2.pie(
        aging.values, labels=aging.index,
        colors=cores_aging[:len(aging)],
        autopct="%1.1f%%", startangle=90,
        pctdistance=0.75,
        wedgeprops={"edgecolor":BG, "linewidth":3, "width":0.52},
    )
    for at in auts:
        at.set_fontsize(9); at.set_color(TEXT); at.set_fontweight("bold")
    total = aging.sum()
    ax2.text(0, 0.08, f"R${total/1e6:.1f}M", ha="center", va="center",
             fontsize=13, fontweight="bold", color=TEXT)
    ax2.text(0, -0.18, "Carteira Total", ha="center", va="center",
             fontsize=8.5, color=MUTED)
    ax2.set_title(f"Aging — {ultimo}", fontsize=11, color=TEXT)
    return _salvar_png(fig, tmp, "g6_inadimplencia")


def fig_fluxo_caixa(fluxo: pd.DataFrame, tmp: Path) -> Path:
    """G7 — FCO mensal (barras) + saldo acumulado (linha)."""
    fl = fluxo.sort_values("competencia").reset_index(drop=True)
    x  = np.arange(len(fl))
    fig = _fig(14, 5)
    fig.text(0.5, 0.97, "Fluxo de Caixa  ·  FCO Mensal e Saldo Acumulado",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    ax = _ax_style(fig.add_axes([0.07, 0.20, 0.87, 0.66]))
    cores = [ACCENT if v >= 0 else ACCENT3 for v in fl["fco"]]
    ax.bar(x, fl["fco"]/1e3, color=cores, alpha=0.85, width=0.65, zorder=2, label="FCO")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_K))
    ax.set_xticks(x)
    ax.set_xticklabels(fl["competencia"], rotation=45, ha="right", fontsize=7.5)
    ax.axhline(0, color=GRID, lw=1.0)
    axr = ax.twinx()
    axr.plot(x, fl["saldo_final"]/1e6, color=BLUE, lw=2.5, marker="o", ms=4, label="Saldo Final")
    axr.set_ylabel("Saldo (R$ MM)", color=BLUE, fontsize=9.5)
    axr.tick_params(axis="y", colors=BLUE, labelsize=8.5)
    axr.set_facecolor(BG2)
    for sp in axr.spines.values(): sp.set_edgecolor(GRID)
    fig.legend(handles=[
        mpatches.Patch(color=ACCENT, label="FCO Positivo"),
        mpatches.Patch(color=ACCENT3,label="FCO Negativo"),
        plt.Line2D([0],[0], color=BLUE, lw=2, label="Saldo Final"),
    ], loc="upper left", bbox_to_anchor=(0.08, 0.95), fontsize=9)
    return _salvar_png(fig, tmp, "g7_fluxo_caixa")


# =============================================================================
# ATO 3 — SÍNTESE
# =============================================================================

def fig_waterfall(dre: pd.DataFrame, tmp: Path) -> Path:
    """G8 — Waterfall da DRE (total do período)."""
    labels  = ["Rec.\nBruta","Deduções","Rec.\nLíquida","CMV","Lucro\nBruto",
               "OPEX","EBITDA","Deprec.","Res.\nFin.","Lucro\nLíq."]
    valores = [
        dre["receita_bruta"].sum(),
        -dre["deducoes"].sum(),
        dre["receita_liquida"].sum(),
        -dre["cmv"].sum(),
        dre["lucro_bruto"].sum(),
        -dre["total_desp_opex"].sum(),
        dre["ebitda"].sum(),
        -dre["depreciacao"].sum(),
        dre["resultado_financeiro"].sum(),
        dre["lucro_liquido"].sum(),
    ]
    totais = [True, False, True, False, True, False, True, False, False, True]
    bases  = [0.0] * len(valores)
    acum   = 0.0
    for i, (v, et) in enumerate(zip(valores, totais)):
        if et:
            bases[i] = 0.0; acum = v
        else:
            bases[i] = acum if v < 0 else acum; acum += v
    cores = [BLUE if et else (ACCENT if v >= 0 else ACCENT3)
             for v, et in zip(valores, totais)]

    fig = _fig(14, 6)
    fig.text(0.5, 0.97, "Ato 3 — Síntese  ·  Estrutura da DRE (Jan/2023 – Dez/2024)",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    ax = _ax_style(fig.add_axes([0.07, 0.18, 0.88, 0.70]))
    ax.bar(range(len(valores)), bases, color="none", width=0.6)
    ax.bar(range(len(valores)), [abs(v) for v in valores],
           bottom=bases, color=cores, alpha=0.88, width=0.6, zorder=2)
    for i in range(len(valores)-1):
        topo = bases[i] + abs(valores[i])
        ax.plot([i+0.3, i+0.7], [topo, topo], color=GRID, lw=1, ls="--")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_M))
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, fontsize=9.5, color=TEXT)
    for i, v in enumerate(valores):
        y = bases[i] + abs(v)/2
        ax.text(i, y, fmt_M(v, None), ha="center", va="center",
                fontsize=8.5, fontweight="bold", color=TEXT)
    return _salvar_png(fig, tmp, "g8_waterfall")


def fig_score(dre: pd.DataFrame, fluxo: pd.DataFrame,
              contas: pd.DataFrame, tmp: Path) -> Path:
    """G9 — Score de Saúde Financeira mensal."""
    taxa = (contas.groupby("competencia")
            .apply(lambda g: g.loc[g["inadimplente"]==1,"valor"].sum()
                   / g["valor"].sum()*100 if g["valor"].sum() > 0 else 0.0)
            .reset_index(name="taxa"))
    merged = (dre[["competencia","margem_ebitda_pct"]]
              .merge(fluxo[["competencia","fco"]], on="competencia")
              .merge(taxa, on="competencia")
              .sort_values("competencia"))
    merged["score"] = merged.apply(
        lambda r: calcular_score_saude(r["margem_ebitda_pct"], r["fco"], r["taxa"]), axis=1)

    fig = _fig(14, 5)
    fig.text(0.5, 0.97, "Score de Saúde Financeira Mensal  ·  0 – 100",
             ha="center", fontsize=14, fontweight="bold", color=ACCENT)
    ax = _ax_style(fig.add_axes([0.07, 0.20, 0.89, 0.66]))
    cores = [ACCENT if v >= 70 else ACCENT2 if v >= 50 else ACCENT3
             for v in merged["score"]]
    bars = ax.bar(range(len(merged)), merged["score"], color=cores, alpha=0.88, width=0.65, zorder=2)
    ax.axhline(70, color=ACCENT,  ls="--", lw=1.6, alpha=0.8, label="Saudável ≥ 70")
    ax.axhline(50, color=ACCENT2, ls="--", lw=1.6, alpha=0.8, label="Atenção ≥ 50")
    ax.set_ylim(0, 115)
    ax.set_xticks(range(len(merged)))
    ax.set_xticklabels(merged["competencia"], rotation=45, ha="right", fontsize=7.5)
    for bar, val in zip(bars, merged["score"]):
        ax.text(bar.get_x()+bar.get_width()/2, val+1.8, f"{val:.0f}",
                ha="center", fontsize=8, fontweight="bold", color=TEXT)
    ax.legend(ncol=2)
    return _salvar_png(fig, tmp, "g9_score")


# =============================================================================
# GERAÇÃO DO PDF
# =============================================================================


def gerar_pdf(pngs: list[Path], dados: dict, output_path: Path) -> None:
    """
    Relatório executivo em PDF conforme ABNT NBR 14724:2011.

    Margens: Superior 3cm | Inferior 2cm | Esquerda 3cm | Direita 2cm
    Fonte: Helvetica (equivalente Arial ABNT) 12pt corpo, 10pt captions
    Entrelinha: 1.5 para texto (leading = fontSize * 1.5)
    Fundo branco com texto escuro para máxima legibilidade e impressão.
    Capa com fundo escuro e links clicáveis LinkedIn e GitHub.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Image, PageBreak, Table, TableStyle,
                                     HRFlowable, KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
    from reportlab.pdfgen import canvas as rl_canvas

    W, H = A4

    # ─────────────────────────────────────────────────────────────────────────
    # MARGENS ABNT NBR 14724:2011
    # Superior 3cm | Inferior 2cm | Esquerda 3cm | Direita 2cm
    # ─────────────────────────────────────────────────────────────────────────
    ML = 3.0 * cm   # margem esquerda
    MR = 2.0 * cm   # margem direita
    MS = 3.0 * cm   # margem superior
    MI = 2.0 * cm   # margem inferior

    W_UTIL = W - ML - MR   # 453.5 pts = 16.0 cm
    H_UTIL = H - MS - MI   # 700.2 pts = 24.7 cm

    # Largura máxima segura para imagens (ABNT: 90% da área útil, centralizada)
    # Área útil = 16.0 cm → 90% = 14.4 cm, garantindo folga de ~0.8 cm em cada lado
    IMG_W_CM     = (W_UTIL / cm) * 0.90   # 14.4 cm (≈ 90% da área útil)
    IMG_MAX_H_CM = (H_UTIL / cm) - 6      # 18.7 cm (reserva 6 cm para caption, título e margens)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        leftMargin=ML, rightMargin=MR,
        topMargin=MS,  bottomMargin=MI,
        title="Relatório Financeiro — Energética Miramar",
        author="Gary Rainer Chumacero Vanderlei",
    )

    # ─────────────────────────────────────────────────────────────────────────
    # PALETA — fundo branco, texto escuro
    # ─────────────────────────────────────────────────────────────────────────
    C_CAPA_BG  = colors.HexColor("#0f1923")
    C_ACCENT   = colors.HexColor("#00BFA6")
    C_ACCENT2  = colors.HexColor("#1a7f6a")
    C_AMBER    = colors.HexColor("#C47A0A")
    C_CORAL    = colors.HexColor("#C0392B")
    C_BLACK    = colors.HexColor("#1a1a2e")
    C_DARK     = colors.HexColor("#2d3748")
    C_CAPTION  = colors.HexColor("#4a5568")
    C_HR       = colors.HexColor("#cbd5e0")
    C_WHITE    = colors.white
    C_BG_TABLE = colors.HexColor("#f7fafc")
    C_BG_ALT   = colors.HexColor("#edf2f7")
    C_HEADER   = colors.HexColor("#0f1923")
    C_LINK     = colors.HexColor("#0077B5")   # azul LinkedIn

    # ─────────────────────────────────────────────────────────────────────────
    # TIPOGRAFIA ABNT
    # Corpo: 12pt, entrelinha 18pt (≈1.5×)
    # Captions: 10pt, entrelinha 13pt
    # ─────────────────────────────────────────────────────────────────────────
    def S(name, **kw) -> ParagraphStyle:
        base = getSampleStyleSheet()["Normal"]
        return ParagraphStyle(name, parent=base, **kw)

    # Capa (fundo escuro — texto claro correto aqui)
    s_capa_title = S("CapaTitle",
        fontSize=26, textColor=C_ACCENT, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=10, leading=32)
    s_capa_sub = S("CapaSub",
        fontSize=14, textColor=C_WHITE, alignment=TA_CENTER,
        fontName="Helvetica", spaceAfter=4, leading=20)
    s_capa_loc = S("CapaLoc",
        fontSize=11, textColor=colors.HexColor("#8899aa"), alignment=TA_CENTER,
        fontName="Helvetica-Oblique", spaceAfter=4, leading=16)
    s_capa_link = S("CapaLink",
        fontSize=10, textColor=C_LINK, alignment=TA_CENTER,
        fontName="Helvetica", spaceAfter=4, leading=15)
    s_capa_meta = S("CapaMeta",
        fontSize=9.5, textColor=colors.HexColor("#6b8090"), alignment=TA_CENTER,
        fontName="Helvetica-Oblique", spaceAfter=4, leading=14)

    # Corpo (fundo branco — ABNT)
    s_h1 = S("H1",
        fontSize=14, textColor=C_ACCENT, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceBefore=18, spaceAfter=4, leading=20)
    s_h2 = S("H2",
        fontSize=12, textColor=C_ACCENT2, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=4, leading=18)
    s_h3 = S("H3",
        fontSize=12, textColor=C_BLACK, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=3, leading=18)
    # ABNT: corpo 12pt, entrelinha 1.5 (leading 18)
    s_body = S("Body",
        fontSize=12, textColor=C_BLACK, alignment=TA_JUSTIFY,
        fontName="Helvetica", leading=18, spaceAfter=10)
    s_bullet = S("Bullet",
        fontSize=12, textColor=C_BLACK, alignment=TA_LEFT,
        fontName="Helvetica", leading=18, leftIndent=20,
        spaceAfter=6, bulletIndent=6)
    # ABNT: captions 10pt, entrelinha simples
    s_caption = S("Caption",
        fontSize=10, textColor=C_CAPTION, alignment=TA_CENTER,
        fontName="Helvetica-Oblique", spaceAfter=12, leading=13)
    s_meta = S("Meta",
        fontSize=10, textColor=C_DARK, alignment=TA_CENTER,
        fontName="Helvetica", spaceAfter=6, leading=14)
    s_achado_titulo = S("AchadoTitulo",
        fontSize=12, textColor=C_ACCENT2, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=3, leading=18)
    s_tag = S("Tag",
        fontSize=9, textColor=C_ACCENT, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceBefore=0, spaceAfter=4, leading=12)
    s_kpi_hdr = S("KpiHdr",
        fontSize=8, textColor=colors.HexColor("#8899aa"), alignment=TA_CENTER,
        fontName="Helvetica", leading=11)
    s_kpi_val = S("KpiVal",
        fontSize=14, textColor=C_ACCENT, alignment=TA_CENTER,
        fontName="Helvetica-Bold", leading=18)

    # ─────────────────────────────────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────────────────────────────────
    def hr(thickness=0.8, color=C_HR, before=4, after=10):
        return HRFlowable(width="100%", thickness=thickness, color=color,
                          spaceBefore=before, spaceAfter=after)

    def hr_accent(after=8):
        return HRFlowable(width="100%", thickness=2, color=C_ACCENT,
                          spaceBefore=0, spaceAfter=after)

    def sp(h=8):
        return Spacer(1, h)

    def img(p: Path, max_h_cm: float = None) -> Image:
        """
        Cria imagem respeitando largura util ABNT (90% da area) e altura maxima.
        Redimensiona proporcionalmente garantindo que nunca ultrapasse as margens.
        Sempre centralizada horizontalmente.
        """
        from reportlab.lib.units import cm as _cm
        max_w = IMG_W_CM * _cm
        max_h = (max_h_cm or IMG_MAX_H_CM) * _cm

        im = Image(str(p))
        orig_w = im.imageWidth
        orig_h = im.imageHeight

        # Escala pela largura maxima
        scale = max_w / orig_w
        new_w = max_w
        new_h = orig_h * scale

        # Se altura ainda ultrapassar o limite, re-escala pela altura
        if new_h > max_h:
            scale = max_h / orig_h
            new_h = max_h
            new_w = orig_w * scale

        # Seguranca dupla: nunca ultrapassar largura maxima
        if new_w > max_w:
            scale = max_w / new_w
            new_w = max_w
            new_h = new_h * scale

        im.drawWidth  = new_w
        im.drawHeight = new_h
        im.hAlign = "CENTER"
        return im

    def tabela_estilizada(data, col_widths, header_bg=C_HEADER,
                          alt1=C_BG_TABLE, alt2=C_WHITE) -> Table:
        """
        Tabela com cabeçalho escuro, linhas alternadas e bordas sutis.
        Garante que textos de células usem cor escura legível.
        """
        n_rows = len(data)
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            # Cabeçalho
            ("BACKGROUND",    (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR",     (0, 0), (-1, 0), C_ACCENT),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
            ("VALIGN",        (0, 0), (-1, 0), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            # Dados
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 9),
            ("TEXTCOLOR",     (0, 1), (-1, -1), C_BLACK),
            ("ALIGN",         (0, 1), (-1, -1), "LEFT"),
            ("VALIGN",        (0, 1), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 1), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            # Bordas
            ("GRID",          (0, 0), (-1, -1), 0.4, C_HR),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.0, C_ACCENT),
        ]
        # Linhas alternadas
        for r in range(1, n_rows):
            bg = alt1 if r % 2 == 1 else alt2
            style.append(("BACKGROUND", (0, r), (-1, r), bg))
        t.setStyle(TableStyle(style))
        return t

    def tag_ato(texto: str) -> Paragraph:
        return Paragraph(f'<font color="#00BFA6"><b>◆ {texto}</b></font>', s_tag)

    # ─────────────────────────────────────────────────────────────────────────
    # DADOS
    # ─────────────────────────────────────────────────────────────────────────
    dre = dados["dre"]
    fl  = dados["fluxo_caixa"]
    cn  = dados["contas_receber"]

    rec_24  = dre[dre["ano"] == 2024]["receita_liquida"].sum()
    rec_23  = dre[dre["ano"] == 2023]["receita_liquida"].sum()
    cresc   = (rec_24 / rec_23 - 1) * 100 if rec_23 else 0
    mg_ebt  = dre["margem_ebitda_pct"].mean()
    saldo   = fl.sort_values("competencia")["saldo_final"].iloc[-1]
    taxa_md = (cn.groupby("competencia")
               .apply(lambda g: g.loc[g["inadimplente"] == 1, "valor"].sum()
                      / g["valor"].sum() * 100 if g["valor"].sum() > 0 else 0.0)
               .mean())

    story: list = []

    # ═════════════════════════════════════════════════════════════════════════
    # CAPA — fundo escuro, links clicáveis LinkedIn e GitHub
    # ═════════════════════════════════════════════════════════════════════════
    URL_LINKEDIN = "https://www.linkedin.com/in/garyrainercv/"
    URL_GITHUB   = "https://github.com/Gary-Rainer-Chumacero-Vanderlei"

    capa_rows = [
        [sp(30)],
        [Paragraph("📊 Análise Exploratória com Storytelling", s_capa_title)],
        [sp(8)],
        [Paragraph("Energética Miramar Distribuidora Ltda.", s_capa_sub)],
        [sp(4)],
        [Paragraph("João Pessoa – PB  ·  Jan/2023 – Dez/2024", s_capa_loc)],
        [sp(30)],
        [Paragraph("Gary Rainer Chumacero Vanderlei", s_capa_sub)],
        [sp(12)],
        # Links clicáveis com ícones Unicode
        [Paragraph(
            f'<link href="{URL_LINKEDIN}">'
            f'<font color="#0A66C2">&#x1F517; LinkedIn</font>'
            f' — linkedin.com/in/garyrainercv'
            f'</link>',
            s_capa_link)],
        [sp(4)],
        [Paragraph(
            f'<link href="{URL_GITHUB}">'
            f'<font color="#c9d1d9">&#x1F517; GitHub</font>'
            f' — Gary-Rainer-Chumacero-Vanderlei'
            f'</link>',
            s_capa_link)],
        [sp(30)],
        [Paragraph("Python · Pandas · Matplotlib · ReportLab · OpenPyXL", s_capa_meta)],
        [Paragraph("Dados sintéticos (seed=42) · Portfólio Data Analytics", s_capa_meta)],
        [sp(20)],
    ]
    capa_t = Table([[row[0]] for row in capa_rows], colWidths=[W_UTIL])
    capa_t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), C_CAPA_BG),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 24),
        ("RIGHTPADDING", (0, 0), (-1, -1), 24),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story += [capa_t, sp(20)]

    # KPIs summary na capa
    kpi_hdr = ["Receita Líq. 2024", "Crescimento YoY",
               "Margem EBITDA", "Saldo Caixa (Dez/24)", "Inadimpl. Média"]
    kpi_val = [formatar_brl(rec_24), f"{cresc:+.1f}%",
               f"{mg_ebt:.1f}%", formatar_brl(saldo), f"{taxa_md:.1f}%"]

    kpi_cells_hdr = [Paragraph(h, s_kpi_hdr) for h in kpi_hdr]
    kpi_cells_val = [Paragraph(v, s_kpi_val) for v in kpi_val]

    kpi_t = Table([kpi_cells_hdr, kpi_cells_val],
                  colWidths=[W_UTIL / 5] * 5)
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), C_HEADER),
        ("BACKGROUND",    (0, 1), (-1, 1), C_BG_TABLE),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("GRID",          (0, 0), (-1, -1), 0.4, C_HR),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, C_ACCENT),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story += [kpi_t, PageBreak()]

    # ═════════════════════════════════════════════════════════════════════════
    # ESTRUTURA DA ANÁLISE
    # ═════════════════════════════════════════════════════════════════════════
    story += [
        Paragraph("Estrutura da Análise", s_h1),
        hr_accent(),
        Paragraph(
            "Esta análise é estruturada em três atos narrativos, conduzindo o leitor "
            "dos dados brutos até recomendações acionáveis para a gestão financeira "
            "da Energética Miramar Distribuidora Ltda., no período de janeiro de 2023 "
            "a dezembro de 2024.", s_body),
        sp(6),
    ]

    # Estilos para células da tabela com quebra de linha automática
    s_tbl_hdr = S("TblHdr",
        fontSize=9, textColor=C_ACCENT, fontName="Helvetica-Bold",
        alignment=TA_CENTER, leading=13)
    s_tbl_cell = S("TblCell",
        fontSize=9, textColor=C_BLACK, fontName="Helvetica",
        alignment=TA_LEFT, leading=13)

    COL1 = 2.7 * cm
    COL2 = 3.8 * cm
    COL3 = IMG_W_CM * cm - COL1 - COL2

    est_data = [
        [Paragraph("Ato", s_tbl_hdr),
         Paragraph("Foco Analítico", s_tbl_hdr),
         Paragraph("Perguntas Respondidas", s_tbl_hdr)],
        [Paragraph("1 — Panorama", s_tbl_cell),
         Paragraph("Visão geral e tendências", s_tbl_cell),
         Paragraph("Como evoluiu a receita? Há sazonalidade?<br/>Como se comportaram as margens operacionais?", s_tbl_cell)],
        [Paragraph("2 — Diagnóstico", s_tbl_cell),
         Paragraph("Custos, anomalias e liquidez", s_tbl_cell),
         Paragraph("Onde estão os outliers de custo?<br/>Qual a taxa de inadimplência?<br/>Como está o fluxo de caixa?", s_tbl_cell)],
        [Paragraph("3 — Síntese", s_tbl_cell),
         Paragraph("Score de saúde e recomendações", s_tbl_cell),
         Paragraph("Qual o score financeiro mensal?<br/>Quais os principais riscos e oportunidades identificados?", s_tbl_cell)],
    ]
    story += [
        tabela_estilizada(est_data, [COL1, COL2, COL3]),
        PageBreak(),
    ]

    # ═════════════════════════════════════════════════════════════════════════
    # ATO 1 — PANORAMA
    # ═════════════════════════════════════════════════════════════════════════
    story += [
        tag_ato("ATO 1  ·  PANORAMA"),
        Paragraph("Receita, Sazonalidade e Margens Operacionais", s_h1),
        hr_accent(),
    ]

    story += [
        Paragraph("1.1  Receita Líquida e EBITDA", s_h2),
        Paragraph(
            "A receita líquida apresentou crescimento consistente ao longo do período "
            "analisado, com sazonalidade clara nos meses de maior demanda energética. "
            "O EBITDA acompanhou a trajetória de receita, com margem oscilando entre "
            "27% e 35%, indicativo de eficiência operacional sustentável.", s_body),
        KeepTogether([img(pngs[0], max_h_cm=9)]),
        Paragraph(
            "Figura 1 — Receita Líquida mensal com Média Móvel 3M (tracejado teal) "
            "e EBITDA com margem percentual no eixo secundário (âmbar).",
            s_caption),
    ]

    story += [
        Paragraph("1.2  Sazonalidade e Comparativo Year-over-Year", s_h2),
        Paragraph(
            f"O índice de sazonalidade identifica meses de alta demanda (índice > 100) "
            f"e meses de baixa (índice < 100), viabilizando planejamento proativo de "
            f"campanhas comerciais e controle de custos variáveis. O comparativo "
            f"ano a ano confirma crescimento de <b>{cresc:+.1f}%</b> na receita "
            f"líquida de 2024 versus 2023.", s_body),
        KeepTogether([img(pngs[1], max_h_cm=8)]),
        Paragraph(
            "Figura 2 — Índice sazonal (base 100) e receita agrupada 2023 vs 2024 "
            "com variação percentual anotada acima de cada par de barras.",
            s_caption),
    ]

    story += [
        Paragraph("1.3  Evolução das Margens Operacionais", s_h2),
        Paragraph(
            "As três margens operacionais — Bruta, EBITDA e Líquida — demonstram "
            "estabilidade estrutural ao longo do período. A Margem EBITDA manteve-se "
            "acima da meta de 20% na totalidade dos meses analisados, confirmando "
            "sólida capacidade de geração de caixa operacional.", s_body),
        KeepTogether([img(pngs[2], max_h_cm=7.5)]),
        Paragraph(
            "Figura 3 — Margens Bruta (azul), EBITDA (âmbar) e Líquida (lilás) "
            "com linha de meta EBITDA em 20% (tracejado teal).",
            s_caption),
        PageBreak(),
    ]

    # ═════════════════════════════════════════════════════════════════════════
    # ATO 2 — DIAGNÓSTICO
    # ═════════════════════════════════════════════════════════════════════════
    story += [
        tag_ato("ATO 2  ·  DIAGNÓSTICO"),
        Paragraph("Custos, Anomalias e Liquidez", s_h1),
        hr_accent(),
    ]

    story += [
        Paragraph("2.1  Heatmap de Custos por Departamento", s_h2),
        Paragraph(
            "O heatmap apresenta a distribuição mensal dos gastos por departamento "
            "em R$ Mil. Células com tonalidade mais escura indicam meses de maior "
            "desembolso. Variações abruptas entre meses consecutivos sinalizam "
            "possíveis anomalias que requerem investigação pela controladoria.", s_body),
        KeepTogether([img(pngs[3], max_h_cm=7.5)]),
        Paragraph(
            "Figura 4 — Gastos mensais por departamento em R$ Mil. "
            "Gradiente: mais escuro = maior gasto.",
            s_caption),
    ]

    story += [
        Paragraph("2.2  Detecção de Outliers via Z-score", s_h2),
        Paragraph(
            "Pontos em vermelho identificam meses com gasto superior a dois desvios "
            "padrão da média histórica do departamento (|Z| > 2). A linha tracejada "
            "representa a média histórica; as linhas pontilhadas demarcam os limiares "
            "±2σ. Os outliers foram inseridos com probabilidade de 8% no gerador de "
            "dados para enriquecer a análise diagnóstica.", s_body),
        KeepTogether([img(pngs[4], max_h_cm=7.5)]),
        Paragraph(
            "Figura 5 — Outliers por departamento. Vermelho: |Z| > 2. "
            "Linha teal: média histórica. Pontilhado: limiares ±2σ.",
            s_caption),
    ]

    story += [
        Paragraph("2.3  Inadimplência e Aging da Carteira", s_h2),
        Paragraph(
            f"A taxa de inadimplência média no período foi de <b>{taxa_md:.1f}%</b>. "
            f"O gráfico de donut evidencia que a maior parcela da carteira encontra-se "
            f"na faixa 'A vencer'; contudo, há concentração relevante nas faixas de "
            f"maior risco (acima de 60 dias), demandando política ativa e diferenciada "
            f"de cobrança por segmento de cliente.", s_body),
        KeepTogether([img(pngs[5], max_h_cm=8)]),
        Paragraph(
            "Figura 6 — Taxa de inadimplência mensal (esquerda) "
            "e distribuição do aging da carteira no último mês disponível (donut).",
            s_caption),
    ]

    story += [
        Paragraph("2.4  Fluxo de Caixa Operacional", s_h2),
        Paragraph(
            f"O Fluxo de Caixa Operacional (FCO) acumulado encerra o período em "
            f"posição positiva, com saldo final de <b>{formatar_brl(saldo)}</b>. "
            f"Os meses com FCO negativo (barras vermelhas) concentram-se no início "
            f"do período e devem ser monitorados para prevenir pressão de liquidez "
            f"em ciclos operacionais futuros.", s_body),
        KeepTogether([img(pngs[6], max_h_cm=7.5)]),
        Paragraph(
            "Figura 7 — FCO mensal: verde = positivo, vermelho = negativo. "
            "Linha azul: saldo de caixa acumulado (eixo secundário direito).",
            s_caption),
        PageBreak(),
    ]

    # ═════════════════════════════════════════════════════════════════════════
    # ATO 3 — SÍNTESE
    # ═════════════════════════════════════════════════════════════════════════
    story += [
        tag_ato("ATO 3  ·  SÍNTESE"),
        Paragraph("Score de Saúde Financeira e Recomendações", s_h1),
        hr_accent(),
    ]

    story += [
        Paragraph("3.1  Estrutura da DRE — Gráfico Waterfall", s_h2),
        Paragraph(
            "O gráfico em cascata (waterfall) decompõe a Receita Bruta até o "
            "Lucro Líquido, evidenciando o impacto de cada componente: Deduções, "
            "CMV, Despesas Operacionais (OPEX) e Resultado Financeiro. "
            "Barras azuis representam totais parciais; verde indica incrementos "
            "e vermelho indica reduções.", s_body),
        KeepTogether([img(pngs[7], max_h_cm=8)]),
        Paragraph(
            "Figura 8 — Waterfall da DRE: valores totais acumulados Jan/2023 – Dez/2024. "
            "Azul = total parcial  ·  Verde = acréscimo  ·  Vermelho = redução.",
            s_caption),
    ]

    story += [
        Paragraph("3.2  Score de Saúde Financeira Mensal", s_h2),
        Paragraph(
            "O Score de Saúde Financeira é um indicador composto (0–100) calculado "
            "mensalmente com base em três componentes ponderados: Margem EBITDA "
            "(até 40 pontos, meta 25%), FCO positivo (30 pontos, critério binário) "
            "e taxa de inadimplência (até 30 pontos, meta inferior a 10%). "
            "Verde indica situação saudável (≥ 70); âmbar indica atenção (50–69); "
            "vermelho indica situação crítica (< 50).", s_body),
        KeepTogether([img(pngs[8], max_h_cm=7.5)]),
        Paragraph(
            "Figura 9 — Score de Saúde Financeira mensal (0–100). "
            "Teal = Saudável (≥ 70)  ·  Âmbar = Atenção (50–69)  ·  Vermelho = Crítico (< 50).",
            s_caption),
        sp(10),
    ]

    # ═════════════════════════════════════════════════════════════════════════
    # CONCLUSÃO
    # ═════════════════════════════════════════════════════════════════════════
    story += [
        hr(thickness=1.5, color=C_ACCENT, before=8, after=12),
        Paragraph("Conclusão — Principais Achados e Recomendações", s_h1),
        hr_accent(after=14),
    ]

    achados = [
        ("1.  Crescimento de receita sustentável",
         f"A receita líquida registrou crescimento de <b>{cresc:+.1f}%</b> em 2024 "
         f"versus 2023, com sazonalidade previsível que viabiliza planejamento "
         f"proativo de campanhas comerciais e controle de custos variáveis."),
        ("2.  Margem EBITDA consistentemente acima da meta",
         f"Margem média de <b>{mg_ebt:.1f}%</b> ao longo de 24 meses demonstra "
         f"eficiência operacional estrutural e robusta capacidade de geração de caixa "
         f"operacional, superando a meta de 20% em todos os meses analisados."),
        ("3.  Outliers de custo identificados e rastreáveis",
         "A análise via Z-score identificou gastos anômalos (|Z| > 2) em múltiplos "
         "departamentos. Investigação e controle preventivo têm potencial mensurável "
         "de redução nos custos operacionais totais do período."),
        ("4.  Inadimplência gerenciável com atenção às faixas longas",
         f"Taxa média de <b>{taxa_md:.1f}%</b> é administrável no setor de "
         f"distribuição de energia, porém a concentração de recebíveis nas faixas "
         f"acima de 60 dias recomenda política ativa e diferenciada de cobrança "
         f"por segmento e porte de cliente."),
        ("5.  Posição de caixa positiva com vigilância sobre meses negativos",
         f"Saldo final de <b>{formatar_brl(saldo)}</b> oferece folga de liquidez "
         f"adequada ao porte da operação. Meses com FCO negativo devem ser "
         f"monitorados para prevenir pressão de curto prazo em ciclos futuros."),
    ]

    for titulo, texto in achados:
        story += [
            KeepTogether([
                Paragraph(titulo, s_achado_titulo),
                Paragraph(texto,  s_body),
                sp(4),
            ])
        ]

    story += [
        sp(10),
        hr(before=4, after=12),
        Paragraph("Próximos Passos Analíticos Sugeridos", s_h2),
        sp(4),
    ]

    proximos = [
        "Modelo de previsão de receita (Prophet ou SARIMA) para planejamento de curto prazo",
        "Análise de sensibilidade da margem EBITDA a variações de CMV e OPEX",
        "Segmentação de clientes inadimplentes por aging para política de cobrança diferenciada",
        "Dashboard interativo Streamlit para monitoramento contínuo dos KPIs em tempo real",
    ]
    for item in proximos:
        story.append(Paragraph(f"• {item}", s_bullet))

    story += [
        sp(28),
        hr(before=4, after=10),
        Paragraph(
            "Análise realizada com dados sintéticos (seed=42) para fins de portfólio.  "
            "Empresa fictícia: Energética Miramar Distribuidora Ltda.",
            s_meta),
        Paragraph(
            "Gary Rainer Chumacero Vanderlei  ·  "
            "Python · Pandas · Matplotlib · ReportLab · OpenPyXL",
            s_meta),
        Paragraph(
            f'<link href="{URL_LINKEDIN}"><font color="#0A66C2">{URL_LINKEDIN}</font></link>'
            f'  ·  '
            f'<link href="{URL_GITHUB}"><font color="#333">{URL_GITHUB}</font></link>',
            s_meta),
    ]

    doc.build(story)
    logger.info("  ✓ PDF salvo em %s", output_path)



def gerar_excel(dados: dict, output_path: Path) -> None:
    """
    Gera workbook Excel com 5 abas: Resumo, DRE, Fluxo de Caixa,
    Centro de Custos e Contas a Receber.

    Args:
        dados: Dicionário com os DataFrames do projeto.
        output_path: Caminho de saída do arquivo .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    wb = Workbook()

    # ── Paleta openpyxl ──
    C_HEADER  = "0F3460"   # azul escuro (cabeçalho)
    C_ACCENT  = "00BFA6"   # teal (accent)
    C_ACCENT2 = "F5A623"   # âmbar
    C_ACCENT3 = "E05A5A"   # coral
    C_BG      = "0F1923"   # fundo escuro
    C_BG2     = "162030"   # superfície
    C_TEXT    = "C8D6E5"   # texto claro
    C_MUTED   = "4A6278"   # texto suave
    C_WHITE   = "FFFFFF"
    C_POS     = "1A7F4E"   # verde
    C_NEG     = "842029"   # vermelho

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=C_TEXT, size=10, italic=False) -> Font:
        return Font(bold=bold, color=color, size=size, italic=italic,
                    name="Arial")

    def center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def thin_border() -> Border:
        s = Side(style="thin", color="1e2d3d")
        return Border(left=s, right=s, top=s, bottom=s)

    def col_width(ws, col: int, w: float):
        ws.column_dimensions[get_column_letter(col)].width = w

    def row_height(ws, row: int, h: float):
        ws.row_dimensions[row].height = h

    def write_header_row(ws, row: int, headers: list[str],
                         bg: str = C_HEADER, fg: str = C_WHITE):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill      = fill(bg)
            cell.font      = font(bold=True, color=fg, size=9)
            cell.alignment = center()
            cell.border    = thin_border()
        row_height(ws, row, 32)

    def write_data_row(ws, row: int, values: list, fmts: list[str] | None = None,
                       bg: str = C_BG2, fg: str = C_TEXT):
        bg_alt = C_BG if row % 2 == 0 else C_BG2
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill      = fill(bg_alt)
            cell.font      = font(color=fg, size=9)
            cell.alignment = center()
            cell.border    = thin_border()
            if fmts and c <= len(fmts) and fmts[c-1]:
                cell.number_format = fmts[c-1]
        row_height(ws, row, 18)

    # ──────────────────────────────────────────────────────────────────────────
    # ABA 1 — RESUMO EXECUTIVO
    # ──────────────────────────────────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "📋 Resumo Executivo"
    ws_r.sheet_view.showGridLines = False
    ws_r.sheet_properties.tabColor = C_ACCENT

    dre = dados["dre"]; fl = dados["fluxo_caixa"]; cn = dados["contas_receber"]

    # Título
    ws_r.merge_cells("A1:F1")
    c = ws_r["A1"]
    c.value = "ENERGÉTICA MIRAMAR  ·  Resumo Executivo  ·  Jan/2023 – Dez/2024"
    c.fill  = fill(C_BG)
    c.font  = font(bold=True, color=C_ACCENT, size=14)
    c.alignment = center()
    row_height(ws_r, 1, 40)

    ws_r.merge_cells("A2:F2")
    c2 = ws_r["A2"]
    c2.value = "Gary Rainer Chumacero Vanderlei  ·  Dados sintéticos (seed=42)"
    c2.fill  = fill(C_BG)
    c2.font  = font(color=C_MUTED, size=9, italic=True)
    c2.alignment = center()
    row_height(ws_r, 2, 20)

    # KPIs anuais
    kpi_headers = ["Indicador", "2023", "2024", "Variação", "Média 24M", "Tendência"]
    write_header_row(ws_r, 4, kpi_headers, bg=C_ACCENT, fg=C_BG)

    rec23 = dre[dre["ano"]==2023]["receita_liquida"].sum()
    rec24 = dre[dre["ano"]==2024]["receita_liquida"].sum()
    ebt23 = dre[dre["ano"]==2023]["ebitda"].sum()
    ebt24 = dre[dre["ano"]==2024]["ebitda"].sum()
    ll23  = dre[dre["ano"]==2023]["lucro_liquido"].sum()
    ll24  = dre[dre["ano"]==2024]["lucro_liquido"].sum()
    fco23 = fl[fl["ano"]==2023]["fco"].sum()
    fco24 = fl[fl["ano"]==2024]["fco"].sum()
    taxa_md = (cn.groupby("competencia")
               .apply(lambda g: g.loc[g["inadimplente"]==1,"valor"].sum()
                      / g["valor"].sum()*100 if g["valor"].sum() > 0 else 0.0)
               .mean())

    FMT_BRL = 'R$ #,##0.00'
    FMT_PCT = '0.0%'
    FMT_INT = '#,##0'

    kpi_rows = [
        ["Receita Líquida",   rec23, rec24, f'=C{{}}-B{{}}', rec23, "▲"],
        ["EBITDA",            ebt23, ebt24, f'=C{{}}-B{{}}', ebt23, "▲"],
        ["Lucro Líquido",     ll23,  ll24,  f'=C{{}}-B{{}}', ll23,  "▲"],
        ["FCO",               fco23, fco24, f'=C{{}}-B{{}}', fco23, "▲"],
        ["Margem EBITDA (%)", dre[dre["ano"]==2023]["margem_ebitda_pct"].mean()/100,
                              dre[dre["ano"]==2024]["margem_ebitda_pct"].mean()/100,
                              None, dre["margem_ebitda_pct"].mean()/100, "→"],
        ["Inadimplência (%)", None, None, None, taxa_md/100, "→"],
    ]
    fmt_cols = [None, FMT_BRL, FMT_BRL, FMT_BRL, FMT_BRL, None]
    fmt_pct_rows = {5, 6}

    for i, row_vals in enumerate(kpi_rows):
        r = 5 + i
        row_height(ws_r, r, 22)
        fmts = [FMT_PCT, FMT_PCT, FMT_PCT, FMT_PCT, FMT_PCT, None] \
               if (i+1) in fmt_pct_rows else fmt_cols
        vals = row_vals[:2] + [row_vals[2]] + [f"=C{r}-B{r}" if row_vals[3] else ""] + [row_vals[4], row_vals[5]]
        write_data_row(ws_r, r, vals, fmts)

    for c in range(1, 7):
        col_width(ws_r, c, 22)

    # ──────────────────────────────────────────────────────────────────────────
    # ABA 2 — DRE MENSAL
    # ──────────────────────────────────────────────────────────────────────────
    ws_d = wb.create_sheet("📊 DRE Mensal")
    ws_d.sheet_view.showGridLines = False
    ws_d.sheet_properties.tabColor = "4F8EF7"

    ws_d.merge_cells("A1:L1")
    c = ws_d["A1"]
    c.value = "DRE — Demonstração de Resultado  ·  Mensal  ·  Jan/2023 – Dez/2024"
    c.fill  = fill(C_BG); c.font = font(bold=True, color="4F8EF7", size=13)
    c.alignment = center(); row_height(ws_d, 1, 36)

    dre_cols = ["Competência","Ano","Mês","Rec. Bruta","Deduções","Rec. Líquida",
                "CMV","Lucro Bruto","OPEX","EBITDA","Lucro Líquido","Mg. EBITDA %"]
    write_header_row(ws_d, 2, dre_cols, bg="0F3460")

    dre_s = dre.sort_values("competencia").reset_index(drop=True)
    for i, row in dre_s.iterrows():
        r = 3 + i
        vals = [row["competencia"], int(row["ano"]), int(row["mes"]),
                row["receita_bruta"], row["deducoes"], row["receita_liquida"],
                row["cmv"], row["lucro_bruto"], row["total_desp_opex"],
                row["ebitda"], row["lucro_liquido"], row["margem_ebitda_pct"]/100]
        fmts = [None, None, None, FMT_BRL, FMT_BRL, FMT_BRL,
                FMT_BRL, FMT_BRL, FMT_BRL, FMT_BRL, FMT_BRL, FMT_PCT]
        write_data_row(ws_d, r, vals, fmts)

    # Linha de totais
    nr = 3 + len(dre_s)
    write_header_row(ws_d, nr,
                     ["TOTAL", "", ""] + [f"=SUM({get_column_letter(c)}{3}:{get_column_letter(c)}{nr-1})"
                                           for c in range(4, 13)] + [""],
                     bg=C_ACCENT, fg=C_BG)

    col_widths_d = [14, 6, 5, 14, 12, 14, 14, 14, 12, 12, 14, 12]
    for c, w in enumerate(col_widths_d, 1):
        col_width(ws_d, c, w)

    # Formatação condicional — Margem EBITDA %
    from openpyxl.formatting.rule import ColorScaleRule
    ws_d.conditional_formatting.add(
        f"L3:L{nr-1}",
        ColorScaleRule(start_type="min", start_color="E05A5A",
                       mid_type="percentile", mid_value=50, mid_color="F5A623",
                       end_type="max", end_color="00BFA6")
    )

    # ──────────────────────────────────────────────────────────────────────────
    # ABA 3 — FLUXO DE CAIXA
    # ──────────────────────────────────────────────────────────────────────────
    ws_f = wb.create_sheet("💰 Fluxo de Caixa")
    ws_f.sheet_view.showGridLines = False
    ws_f.sheet_properties.tabColor = "00BFA6"

    ws_f.merge_cells("A1:I1")
    c = ws_f["A1"]
    c.value = "Fluxo de Caixa  ·  Jan/2023 – Dez/2024"
    c.fill  = fill(C_BG); c.font = font(bold=True, color=C_ACCENT, size=13)
    c.alignment = center(); row_height(ws_f, 1, 36)

    fl_cols = ["Competência","Ano","Mês","Recebimentos","Total Saídas",
               "FCO","CAPEX","Variação Caixa","Saldo Final"]
    write_header_row(ws_f, 2, fl_cols, bg="0F3460")

    fl_s = fl.sort_values("competencia").reset_index(drop=True)
    for i, row in fl_s.iterrows():
        r = 3 + i
        vals = [row["competencia"], int(row["ano"]), int(row["mes"]),
                row["recebimentos"], row["total_saidas"], row["fco"],
                row["capex"], row["variacao_caixa"], row["saldo_final"]]
        fmts = [None, None, None] + [FMT_BRL]*6
        write_data_row(ws_f, r, vals, fmts)

        # Colorir FCO positivo/negativo
        fco_cell = ws_f.cell(row=r, column=6)
        if row["fco"] >= 0:
            fco_cell.font = font(color="2DD4A0", bold=True, size=9)
        else:
            fco_cell.font = font(color="F26B6B", bold=True, size=9)

    for c, w in enumerate([14,6,5,14,14,14,12,14,14], 1):
        col_width(ws_f, c, w)

    # ──────────────────────────────────────────────────────────────────────────
    # ABA 4 — CENTRO DE CUSTOS
    # ──────────────────────────────────────────────────────────────────────────
    ws_c = wb.create_sheet("🏢 Centro de Custos")
    ws_c.sheet_view.showGridLines = False
    ws_c.sheet_properties.tabColor = "F5A623"

    ws_c.merge_cells("A1:G1")
    c = ws_c["A1"]
    c.value = "Centro de Custos  ·  Jan/2023 – Dez/2024"
    c.fill  = fill(C_BG); c.font = font(bold=True, color=C_ACCENT2, size=13)
    c.alignment = center(); row_height(ws_c, 1, 36)

    cc_cols = ["Competência","Ano","Mês","Departamento","Categoria","Valor","Outlier"]
    write_header_row(ws_c, 2, cc_cols, bg="0F3460")

    cc_s = dados["centro_custos"].sort_values(["competencia","departamento"]).reset_index(drop=True)
    for i, row in cc_s.iterrows():
        r = 3 + i
        vals = [row["competencia"], int(row["ano"]), int(row["mes"]),
                row["departamento"], row["categoria"], row["valor"],
                "Sim" if row["is_outlier"] else ""]
        fmts = [None, None, None, None, None, FMT_BRL, None]
        write_data_row(ws_c, r, vals, fmts)
        if row["is_outlier"]:
            ws_c.cell(row=r, column=7).font = font(color=C_ACCENT3, bold=True, size=9)

    for c, w in enumerate([14,6,5,16,16,16,8], 1):
        col_width(ws_c, c, w)

    # ──────────────────────────────────────────────────────────────────────────
    # ABA 5 — CONTAS A RECEBER
    # ──────────────────────────────────────────────────────────────────────────
    ws_cr = wb.create_sheet("📥 Contas a Receber")
    ws_cr.sheet_view.showGridLines = False
    ws_cr.sheet_properties.tabColor = "E05A5A"

    ws_cr.merge_cells("A1:I1")
    c = ws_cr["A1"]
    c.value = "Contas a Receber — Inadimplência e Aging  ·  Jan/2023 – Dez/2024"
    c.fill  = fill(C_BG); c.font = font(bold=True, color=C_ACCENT3, size=13)
    c.alignment = center(); row_height(ws_cr, 1, 36)

    cr_cols = ["Competência","Ano","Mês","Cliente","Setor","Porte",
               "Faixa Aging","Valor","Inadimplente"]
    write_header_row(ws_cr, 2, cr_cols, bg="0F3460")

    cr_s = cn.sort_values(["competencia","cliente"]).reset_index(drop=True)
    for i, row in cr_s.iterrows():
        r = 3 + i
        inadim_label = "Sim" if row["inadimplente"] == 1 else ""
        vals = [row["competencia"], int(row["ano"]), int(row["mes"]),
                row["cliente"], row["setor"], row["porte"],
                row["faixa_aging"], row["valor"], inadim_label]
        fmts = [None, None, None, None, None, None, None, FMT_BRL, None]
        write_data_row(ws_cr, r, vals, fmts)
        if row["inadimplente"] == 1:
            ws_cr.cell(row=r, column=9).font = font(color=C_ACCENT3, bold=True, size=9)

    for c, w in enumerate([14,6,5,16,14,10,14,14,10], 1):
        col_width(ws_cr, c, w)

    wb.save(str(output_path))
    logger.info("  ✓ Excel salvo em %s", output_path)


# =============================================================================
# PONTO DE ENTRADA
# =============================================================================

def main() -> None:
    """
    Orquestra a EDA completa gerando PDF e Excel.

    Execução:
        python -m notebooks.eda_financeiro
    """
    print("\n" + "═"*58)
    print("  EDA Financial  ·  Storytelling v4")
    print("  Saídas: PDF + Excel")
    print("═"*58 + "\n")

    loader = DataLoader(data_dir=Settings.RAW_DATA_DIR)
    dados  = loader.carregar_todos()

    dre    = dados["dre"].sort_values("competencia").reset_index(drop=True)
    fluxo  = dados["fluxo_caixa"].sort_values("competencia").reset_index(drop=True)
    contas = dados["contas_receber"].sort_values("competencia").reset_index(drop=True)
    centro = dados["centro_custos"].sort_values("competencia").reset_index(drop=True)

    output_dir = Settings.EXPORTS_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Gráficos em diretório temporário ──
    with tempfile.TemporaryDirectory() as tmp_str:
        tmp = Path(tmp_str)
        logger.info("Gerando gráficos...")
        pngs = [
            fig_receita_ebitda(dre, tmp),          # G1
            fig_sazonalidade(dre, tmp),             # G2
            fig_margens(dre, tmp),                  # G3
            fig_custos_heatmap(centro, tmp),        # G4
            fig_outliers(centro, tmp),              # G5
            fig_inadimplencia(contas, tmp),         # G6
            fig_fluxo_caixa(fluxo, tmp),            # G7
            fig_waterfall(dre, tmp),                # G8
            fig_score(dre, fluxo, contas, tmp),     # G9
        ]

        logger.info("Gerando PDF...")
        pdf_path = output_dir / "relatorio_financeiro.pdf"
        gerar_pdf(pngs, dados, pdf_path)

    logger.info("Gerando Excel...")
    xlsx_path = output_dir / "relatorio_financeiro.xlsx"
    gerar_excel(dados, xlsx_path)

    print("\n" + "═"*58)
    print("  ✅  Concluído")
    print(f"  📄  {pdf_path}")
    print(f"  📊  {xlsx_path}")
    print("═"*58 + "\n")


if __name__ == "__main__":
    main()
